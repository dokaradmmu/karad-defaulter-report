import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

# ── Constants ────────────────────────────────────────────────────────────────
EXCLUDED_OFFICES = {"Shenawadi B.O", "Yeralwadi B.O"}
KPI = 0.90

NAVY        = "1F3864"
LIGHT_BLUE  = "DCE6F1"
WHITE       = "FFFFFF"
RED         = "C00000"

SUB_DIV_ORDER = ["SDIP Karad East", "SDIP Vaduj", "ASP Karad West"]
SHEET_NAMES   = ["Karad East", "Vaduj", "Karad West"]
TAB_COLOURS   = ["1F3864", "375623", "7F3F00"]

COL_WIDTHS = [7, 22, 26, 28, 11, 22, 20, 18]

THIN = Side(style="thin", color="AAAAAA")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Helpers ──────────────────────────────────────────────────────────────────
def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _align(h="center", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _apply_border(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = ALL_BORDER

def _merge_write(ws, row, col_start, col_end, value, font, fill, align, height=None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = align
    cell.border = ALL_BORDER
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).border = ALL_BORDER
    if height:
        ws.row_dimensions[row].height = height

# ── Master file loader ────────────────────────────────────────────────────────
def load_master(file_obj):
    df = pd.read_excel(file_obj, dtype=str)
    df.columns = df.columns.str.strip()
    for col in ["Sub Division Name", "Sub Office Name", "Office ID", "Office Name", "Office Type Code"]:
        df[col] = df[col].str.strip()
    df = df[~df["Office Name"].isin(EXCLUDED_OFFICES)]
    return df[["Sub Division Name", "Sub Office Name", "Office ID", "Office Name", "Office Type Code"]]

# ── CSV loaders ───────────────────────────────────────────────────────────────
def load_dp_csv(file_obj):
    df = pd.read_csv(file_obj, dtype=str)
    df.columns = df.columns.str.strip()
    df["office-id"] = df["office-id"].str.strip()
    for col in ["invoice-count", "delivery-count", "return-count", "redirection-count"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df[["office-id", "invoice-count", "delivery-count", "return-count", "redirection-count"]]

def load_dss_csv(file_obj):
    df = pd.read_csv(file_obj, dtype=str)
    df.columns = df.columns.str.strip()
    # Drop summary row where office_id == 0
    df = df[df["office_id"].str.strip() != "0"]
    df["office_id"] = df["office_id"].str.strip()
    for col in ["total_pdm_art_count", "total_dss_art_count"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df[["office_id", "total_pdm_art_count", "total_dss_art_count"]]

# ── DP calculation ────────────────────────────────────────────────────────────
def calc_dp(df_csv):
    result = {}
    for _, row in df_csv.iterrows():
        inv = row["invoice-count"]
        if inv > 0:
            pct = (row["delivery-count"] + row["return-count"] + row["redirection-count"]) / inv
            result[row["office-id"]] = {"pct": pct, "invoice": inv}
    return result

# ── DSS calculation ───────────────────────────────────────────────────────────
def calc_dss(df_csv):
    result = {}
    for _, row in df_csv.iterrows():
        pdm = row["total_pdm_art_count"]
        dss = row["total_dss_art_count"]
        if pdm > 0:
            result[row["office_id"]] = {"pct": dss / pdm, "pdm": pdm, "dss": dss}
        # pdm == 0 → absent from result (blank)
    return result

# ── Sort master offices ───────────────────────────────────────────────────────
def sort_offices(df_master, subdiv_name):
    sub = df_master[df_master["Sub Division Name"] == subdiv_name].copy()
    type_order = {"HPO": 0, "SPO": 1, "BPO": 2}
    sub["_type_rank"] = sub["Office Type Code"].map(type_order).fillna(9)
    sub = sub.sort_values(["Sub Office Name", "_type_rank", "Office Name"])
    return sub.drop(columns=["_type_rank"])

# ── Write block ───────────────────────────────────────────────────────────────
def _write_block(ws, start_row, offices_df, cum_data, daily_data,
                 block_type, title_text, col_f_header, report_date_str):
    """
    block_type: 'dp' or 'dss'
    Returns next free row after block.
    """
    # Filter defaulters (cumulative < 90%)
    defaulters = []
    for _, off in offices_df.iterrows():
        oid = off["Office ID"]
        cum = cum_data.get(oid)
        if cum and cum["pct"] < KPI:
            defaulters.append(off)

    row = start_row

    # Title row
    _merge_write(
        ws, row, 1, 8,
        title_text,
        _font(bold=True, color="000000", size=18),
        _fill(WHITE),
        _align("center"),
        height=51
    )
    row += 1

    # Header row
    headers = [
        "Sr. No.",
        "Sub Division Name",
        "Sub Office Name",
        "Office Name",
        "Office Type",
        col_f_header,
        f"Total Articles\nInvoiced {report_date_str}",
        f"{'Delivery Productivity' if block_type == 'dp' else 'DSS Usage'} %\n{report_date_str}",
    ]
    for c, (hdr, w) in enumerate(zip(headers, COL_WIDTHS), start=1):
        cell = ws.cell(row=row, column=c, value=hdr)
        cell.font = _font(bold=True, color="FFFFFF", size=10)
        cell.fill = _fill(NAVY)
        cell.alignment = _align("center")
        cell.border = ALL_BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 40
    row += 1

    if not defaulters:
        _merge_write(
            ws, row, 1, 8,
            "No defaulter offices found for this Sub Division.",
            _font(size=10),
            _fill(WHITE),
            _align("center"),
            height=18
        )
        row += 1
        return row

    for sr, off in enumerate(defaulters, start=1):
        oid = off["Office ID"]
        bg = WHITE if sr % 2 == 1 else LIGHT_BLUE

        cum = cum_data.get(oid)
        daily = daily_data.get(oid)

        # Col A — Sr No
        c = ws.cell(row=row, column=1, value=sr)
        c.font = _font(size=10); c.fill = _fill(bg); c.alignment = _align("center"); c.border = ALL_BORDER

        # Cols B-E — names
        for ci, val in enumerate([
            off["Sub Division Name"], off["Sub Office Name"],
            off["Office Name"], off["Office Type Code"]
        ], start=2):
            halign = "left" if ci in (2, 3, 4) else "center"
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = _font(size=10); cell.fill = _fill(bg)
            cell.alignment = _align(halign); cell.border = ALL_BORDER

        # Col F — cumulative % (always red bold, it's a defaulter by definition)
        f_cell = ws.cell(row=row, column=6, value=cum["pct"])
        f_cell.font = _font(bold=True, color=RED, size=10)
        f_cell.fill = _fill(bg)
        f_cell.number_format = "0.00%"
        f_cell.alignment = _align("center"); f_cell.border = ALL_BORDER

        # Col G — daily invoice/pdm count
        if block_type == "dp":
            g_val = int(daily["invoice"]) if daily else None
        else:
            g_val = int(daily["pdm"]) if daily and daily["pdm"] > 0 else None
        g_cell = ws.cell(row=row, column=7, value=g_val)
        g_cell.font = _font(size=10); g_cell.fill = _fill(bg)
        g_cell.number_format = "#,##0"
        g_cell.alignment = _align("center"); g_cell.border = ALL_BORDER

        # Col H — daily %
        h_cell = ws.cell(row=row, column=8)
        h_cell.fill = _fill(bg); h_cell.alignment = _align("center"); h_cell.border = ALL_BORDER

        if block_type == "dp":
            if daily:
                h_val = daily["pct"]
                h_cell.value = h_val
                h_cell.number_format = "0.00%"
                is_red = h_val < KPI
                h_cell.font = _font(bold=is_red, color=RED if is_red else "000000", size=10)
            else:
                h_cell.value = None
                h_cell.font = _font(size=10)
        else:  # dss
            if daily is None or daily["pdm"] == 0:
                h_cell.value = None
                h_cell.font = _font(size=10)
            elif daily["dss"] == 0:
                h_cell.value = 0.0
                h_cell.number_format = "0.00%"
                h_cell.font = _font(bold=True, color=RED, size=10)
            else:
                h_val = daily["pct"]
                h_cell.value = h_val
                h_cell.number_format = "0.00%"
                is_red = h_val < KPI
                h_cell.font = _font(bold=is_red, color=RED if is_red else "000000", size=10)

        ws.row_dimensions[row].height = 18
        row += 1

    return row

# ── Main builder ──────────────────────────────────────────────────────────────
def build_report(
    master_file, dp_cum_file, dp_daily_file, dss_cum_file, dss_daily_file,
    cum_from: date, cum_to: date, daily_date: date
) -> tuple[bytes, dict, list]:
    """
    Returns: (xlsx_bytes, summary_dict, unmatched_list)
    summary_dict = {sheet_name: {"dp": int, "dss": int}}
    unmatched_list = list of office IDs not found in master
    """
    report_date = daily_date + timedelta(days=1)

    # Format strings
    fmt = lambda d: d.strftime("%d.%m.%Y")
    fmt_file = lambda d: d.strftime("%d_%m_%Y")

    cum_from_s = fmt(cum_from)
    cum_to_s   = fmt(cum_to)
    daily_s    = fmt(daily_date)
    report_s   = fmt(report_date)

    # Load data
    master  = load_master(master_file)
    dp_cum  = calc_dp(load_dp_csv(dp_cum_file))
    dp_day  = calc_dp(load_dp_csv(dp_daily_file))
    dss_cum = calc_dss(load_dss_csv(dss_cum_file))
    dss_day = calc_dss(load_dss_csv(dss_daily_file))

    # Unmatched IDs
    master_ids = set(master["Office ID"])
    all_csv_ids = set(dp_cum) | set(dp_day) | set(dss_cum) | set(dss_day)
    unmatched = sorted(all_csv_ids - master_ids)

    wb = Workbook()
    wb.remove(wb.active)

    summary = {}

    for subdiv, sheet_name, tab_col in zip(SUB_DIV_ORDER, SHEET_NAMES, TAB_COLOURS):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab_col
        ws.sheet_view.showGridLines = False

        offices = sort_offices(master, subdiv)

        # Count defaulters for summary
        dp_def_count  = sum(1 for _, o in offices.iterrows() if (d := dp_cum.get(o["Office ID"])) and d["pct"] < KPI)
        dss_def_count = sum(1 for _, o in offices.iterrows() if (d := dss_cum.get(o["Office ID"])) and d["pct"] < KPI)
        summary[sheet_name] = {"dp": dp_def_count, "dss": dss_def_count}

        # Block 1 — DP
        dp_title = (
            f"Delivery Productivity Report {report_s} "
            f"( या यादीतील ऑफिसेस चा एकूण Delivery Productivity % "
            f"०१ तारखेपासून काल पर्यंत ९०% पेक्षा कमी आहे)"
        )
        col_f_dp = f"Delivery Productivity %\n{cum_from_s} to {cum_to_s}"
        next_row = _write_block(ws, 1, offices, dp_cum, dp_day, "dp", dp_title, col_f_dp, daily_s)

        # Two blank rows
        next_row += 2

        # Block 2 — DSS
        dss_title = (
            f"DSS Usage Defaulter Offices {report_s} "
            f"( या यादीतील ऑफिसेस चा एकूण DSS % "
            f"०१ तारखेपासून काल पर्यंत ९०% पेक्षा कमी आहे)"
        )
        col_f_dss = f"DSS Usage %\n{cum_from_s} to {cum_to_s}"
        _write_block(ws, next_row, offices, dss_cum, dss_day, "dss", dss_title, col_f_dss, daily_s)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read(), summary, unmatched, fmt_file(report_date)
